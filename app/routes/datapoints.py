# app/routes/datapoints.py
import csv
import io
import uuid

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app.core.security import get_current_user
from app.database import get_db
from app.models.datapoint import DataPoint
from app.schemas.datapoint import (
    DataPointCreateSchema,
    DataPointListResponse,
    DataPointResponse,
    DataPointUpdateSchema,
    ImportReportSchema,
)
from app.schemas.user import MessageResponse

router = APIRouter(prefix="/api/points", tags=["DataPoints"])

IMPORT_COLUMNS = ["name", "lat", "lng", "zone", "type", "status", "score", "revenue"]
EXPORT_COLUMNS = [
    "id", "name", "lat", "lng", "zone", "type", "status", "score",
    "revenue", "created_by", "created_at", "updated_at",
]


def _apply_filters(query, q, type, zone, status, score_min):
    if q:
        query = query.filter(DataPoint.name.ilike(f"%{q}%"))
    if type:
        query = query.filter(DataPoint.type == type)
    if zone:
        query = query.filter(DataPoint.zone == zone)
    if status:
        query = query.filter(DataPoint.status == status)
    if score_min is not None:
        query = query.filter(DataPoint.score >= score_min)
    return query


def _row_to_schema(row: dict) -> DataPointCreateSchema:
    payload = {
        "name": (row.get("name") or "").strip(),
        "lat": float(row["lat"]),
        "lng": float(row["lng"]),
        "zone": (row.get("zone") or "").strip(),
        "type": (row.get("type") or "").strip(),
    }
    if row.get("status"):
        payload["status"] = str(row["status"]).strip()
    if row.get("score") not in (None, ""):
        payload["score"] = int(row["score"])
    if row.get("revenue") not in (None, ""):
        payload["revenue"] = float(row["revenue"])
    return DataPointCreateSchema(**payload)


def _read_rows(filename: str, content: bytes) -> list[dict]:
    if filename.endswith(".xlsx"):
        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        return [dict(zip(headers, row)) for row in rows_iter]

    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))

    raise HTTPException(status_code=400, detail="Format de fichier non supporté (attendu .csv ou .xlsx)")


@router.get("", response_model=DataPointListResponse)
def list_points(
    q: str | None = None,
    type: str | None = None,
    zone: str | None = None,
    status: str | None = None,
    score_min: int | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    query = _apply_filters(db.query(DataPoint), q, type, zone, status, score_min)

    total = query.count()
    items = (
        query.order_by(DataPoint.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return DataPointListResponse(items=items, total=total, page=page, page_size=page_size)


@router.post("", response_model=DataPointResponse, status_code=201)
def create_point(
    body: DataPointCreateSchema,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    point = DataPoint(**body.model_dump(), created_by=current_user.id)
    db.add(point)
    db.commit()
    db.refresh(point)
    return point


@router.post("/import", response_model=ImportReportSchema)
async def import_points(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    content = await file.read()
    rows = _read_rows(file.filename or "", content)

    valid_points = []
    errors = []
    for line_number, row in enumerate(rows, start=2):
        try:
            data = _row_to_schema(row)
        except (ValueError, ValidationError, KeyError) as exc:
            errors.append({"line": line_number, "reason": str(exc)})
            continue
        valid_points.append(DataPoint(**data.model_dump(), created_by=None))

    # Un seul commit final : une ligne invalide ne doit ni interrompre le traitement
    # des lignes suivantes, ni faire échouer partiellement les lignes déjà validées.
    if valid_points:
        db.add_all(valid_points)
        db.commit()

    return ImportReportSchema(inserted=len(valid_points), rejected=len(errors), errors=errors)


@router.get("/export")
def export_points(
    q: str | None = None,
    type: str | None = None,
    zone: str | None = None,
    status: str | None = None,
    score_min: int | None = None,
    format: str = Query("csv"),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    if format not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="Format invalide, attendu 'csv' ou 'xlsx'")

    points = _apply_filters(db.query(DataPoint), q, type, zone, status, score_min).order_by(
        DataPoint.created_at.desc()
    ).all()

    def _row_values(p: DataPoint) -> list:
        return [
            str(p.id), p.name, p.lat, p.lng, p.zone, p.type, p.status, p.score,
            p.revenue, p.created_by, p.created_at.isoformat(), p.updated_at.isoformat(),
        ]

    if format == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(EXPORT_COLUMNS)
        for p in points:
            writer.writerow(_row_values(p))
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="datapoints_export.csv"'},
        )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(EXPORT_COLUMNS)
    for p in points:
        sheet.append(_row_values(p))
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="datapoints_export.xlsx"'},
    )


@router.get("/{point_id}", response_model=DataPointResponse)
def get_point(point_id: uuid.UUID, db: Session = Depends(get_db), _user=Depends(get_current_user)):
    point = db.query(DataPoint).filter(DataPoint.id == point_id).first()
    if not point:
        raise HTTPException(status_code=404, detail="Point introuvable")
    return point


@router.put("/{point_id}", response_model=DataPointResponse)
def update_point(
    point_id: uuid.UUID,
    body: DataPointUpdateSchema,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    point = db.query(DataPoint).filter(DataPoint.id == point_id).first()
    if not point:
        raise HTTPException(status_code=404, detail="Point introuvable")

    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(point, field, value)

    db.commit()
    db.refresh(point)
    return point


@router.delete("/{point_id}", response_model=MessageResponse)
def delete_point(point_id: uuid.UUID, db: Session = Depends(get_db), _user=Depends(get_current_user)):
    point = db.query(DataPoint).filter(DataPoint.id == point_id).first()
    if not point:
        raise HTTPException(status_code=404, detail="Point introuvable")

    # Suppression réelle : aucune table n'a encore de FK vers data_points.id.
    # À revoir en soft delete si un futur module (Zones, Reports) en ajoute une.
    db.delete(point)
    db.commit()
    return {"message": "Point supprimé avec succès"}
